# -*- coding: utf-8 -*-
"""
ExtratoExtractor - Camada EXTRACT do ETL
Responsável apenas por ler e limpar arquivos Excel, retornando dados brutos.
"""
import os
from typing import List, Dict, Any
from datetime import datetime
import pyexcel as p
from openpyxl import load_workbook


class ExtratoExtractor:
    """
    Extrator de dados de extratos bancários.
    
    Responsabilidades:
    - Ler arquivos .xls/.xlsx
    - Limpar formatos e mesclagens
    - Remover colunas vazias
    - Retornar dados brutos (sem transformações)
    """
    
    @staticmethod
    def extract_from_files(file_paths: List[str]) -> List[Dict[str, Any]]:
        """
        Extrai dados brutos de múltiplos arquivos Excel.
        
        Args:
            file_paths: Lista de caminhos dos arquivos a processar
            
        Returns:
            Lista de dicionários com dados brutos:
            [
                {
                    "data_hora": datetime,
                    "categoria": str,
                    "transacao": str,
                    "descricao": str,
                    "valor": float
                },
                ...
            ]
        """
        print(f"📂 [EXTRACTOR] Extraindo dados de {len(file_paths)} arquivo(s)...")
        
        all_records = []
        
        for idx, file_path in enumerate(file_paths, 1):
            filename = os.path.basename(file_path)
            print(f"📄 [EXTRACTOR] Arquivo {idx}/{len(file_paths)}: {filename}")
            
            records = ExtratoExtractor._extract_from_single_file(file_path)
            all_records.extend(records)
            print(f"   ✅ Extraídos: {len(records)} registros")
        
        print(f"✅ [EXTRACTOR] Total extraído: {len(all_records)} registros")
        return all_records
    
    @staticmethod
    def _extract_from_single_file(file_path: str) -> List[Dict[str, Any]]:
        """
        Extrai dados brutos de um único arquivo Excel.
        
        Args:
            file_path: Caminho do arquivo
            
        Returns:
            Lista de dicionários com dados brutos
        """
        # Converte .xls para .xlsx se necessário
        if file_path.endswith('.xls'):
            temp_xlsx = file_path.replace('.xls', '_temp.xlsx')
            p.save_book_as(file_name=file_path, dest_file_name=temp_xlsx)
            file_path = temp_xlsx
        
        # Carrega workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        # Limpa formatação
        ExtratoExtractor._clean_formatting(sheet)

        # Remove colunas desnecessárias (A, E, F, H, I, J, M, N)
        ExtratoExtractor._remove_columns(sheet, ['A', 'E', 'F', 'H', 'I', 'J', 'M', 'N'])

        # Remove linhas inválidas
        ExtratoExtractor._remove_invalid_rows(sheet)

        # Remove cabeçalhos do banco (linhas 1-10)
        sheet.delete_rows(1, 10)

        # Remove linha de cabeçalho (antiga linha 11, agora linha 1)
        # Linha 1 agora contém: "Data e hora", "Categoria", "Transação", "Descrição", "Valor"
        sheet.delete_rows(1, 1)

        # Extrai dados brutos
        records = ExtratoExtractor._extract_records(sheet)
        
        # Limpa arquivo temporário se foi criado
        if file_path.endswith('_temp.xlsx'):
            os.remove(file_path)
        
        return records
    
    @staticmethod
    def _clean_formatting(sheet):
        """Remove toda formatação das células."""
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = None
                cell.fill = None
                cell.border = None
                cell.alignment = None
    
    @staticmethod
    def _remove_columns(sheet, column_letters: List[str]):
        """Remove colunas específicas."""
        column_indices = [ExtratoExtractor._column_index_from_letter(letter) for letter in column_letters]
        column_indices.sort(reverse=True)  # Remove da última para a primeira
        
        for col_index in column_indices:
            sheet.delete_cols(col_index)
    
    @staticmethod
    def _remove_invalid_rows(sheet):
        """
        Remove linhas inválidas:
        - Sem valor na coluna 5 (Valor)
        - Sem data válida na coluna 1 (Data e hora)
        - Sem categoria na coluna 2 (Categoria)
        - Com descrição "Saldo Diário" na coluna 4 (Descrição)
        """
        deleted_count = 0
        for row in range(sheet.max_row, 11, -1):
            col1_value = sheet.cell(row=row, column=1).value  # Data e hora
            col2_value = sheet.cell(row=row, column=2).value  # Categoria
            col4_value = sheet.cell(row=row, column=4).value  # Descrição
            col5_value = sheet.cell(row=row, column=5).value  # Valor

            # Remove linha sem valor na coluna 5
            if not col5_value:
                sheet.delete_rows(row)
                deleted_count += 1
                continue

            # Remove linha sem data válida na coluna 1
            if not ExtratoExtractor._is_date(col1_value):
                sheet.delete_rows(row)
                deleted_count += 1
                continue

            # Remove linha sem categoria na coluna 2
            if not col2_value:
                sheet.delete_rows(row)
                deleted_count += 1
                continue

            # Remove linha com descrição "Saldo Diário"
            if col4_value and str(col4_value).strip().lower() == "saldo diário":
                sheet.delete_rows(row)
                deleted_count += 1
                continue
    
    @staticmethod
    def _extract_records(sheet) -> List[Dict[str, Any]]:
        """
        Extrai registros brutos do sheet.

        Estrutura esperada após limpeza:
        Coluna 1: Data e hora
        Coluna 2: Categoria
        Coluna 3: Transação
        Coluna 4: Descrição
        Coluna 5: Valor

        Nota: Após delete_rows(1, 10), a linha 1 já é o primeiro registro de dados,
        não há mais cabeçalho.
        """
        records = []

        # Começa da linha 1 porque já deletamos os cabeçalhos do banco (linhas 1-10)
        for row in range(1, sheet.max_row + 1):
            record = {
                "data_hora": sheet.cell(row, 1).value,
                "categoria": sheet.cell(row, 2).value,
                "transacao": sheet.cell(row, 3).value,
                "descricao": sheet.cell(row, 4).value,
                "valor": sheet.cell(row, 5).value
            }
            records.append(record)

        return records

    @staticmethod
    def _column_index_from_letter(letter: str) -> int:
        """Converte letra da coluna (A, B, C...) para índice (1, 2, 3...)."""
        return ord(letter.upper()) - ord('A') + 1

    @staticmethod
    def _is_date(value) -> bool:
        """Verifica se o valor é uma data válida."""
        if isinstance(value, datetime):
            return True
        if isinstance(value, str):
            # Tenta formato com hora: "19/11/2025 00:31"
            try:
                datetime.strptime(value, "%d/%m/%Y %H:%M")
                return True
            except:
                pass
            # Tenta formato sem hora: "19/11/2025"
            try:
                datetime.strptime(value, "%d/%m/%Y")
                return True
            except:
                pass
        return False

